import pandas as pd
import requests
from datetime import datetime, timedelta
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os
import json
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def convert_date_format(date_str):
    """Convert date from YYYYMMDD to YYYY-MM-DD format."""
    try:
        return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")
    except ValueError:
        return date_str


def generate_kepco_daily_report():
    """Generate daily total power usage report (kepcoDailyData logic)."""
    try:
        # Set date to yesterday
        today = datetime.today() - timedelta(days=1)
        date = today.strftime("%Y%m%d")  # Format as YYYYMMDD

        # Input CSV file path
        # input_csv = "kepcolist_gg.csv"
        input_csv = "/home/nit/kepco_report/kepcolist_gg.csv"


        # API URL and static parameters
        url = "https://opm.kepco.co.kr:11080/OpenAPI/getDayLpData.do"
        service_key = "bpb89eyd7bg430vckh8t"

        # Read customer numbers from CSV
        customer_data = pd.read_csv(input_csv, dtype=str)

        # Prepare a list to store the results
        results = []

        for _, row in customer_data.iterrows():
            cust_no = row.get("고객번호")
            if not cust_no:
                continue
            bonbu = row.get("본부명")
            center = row.get("센터")
            team = row.get("팀")
            guksa = row.get("국사")

            # API call
            params = {
                "custNo": cust_no,
                "date": date,
                "serviceKey": service_key,
                "returnType": "02",
            }
            response = requests.get(url, params=params)

            if response.status_code == 200:
                data = response.json()
                day_lp_data = data.get("dayLpDataInfoList", [])

                if day_lp_data:
                    total_power = 0
                    for record in day_lp_data:
                        total_power += sum(
                            value
                            for key, value in record.items()
                            if key.startswith("pwr_qty")
                            and isinstance(value, (int, float))
                        )
                    results.append(
                        {
                            "Customer Number": cust_no,
                            "Date": convert_date_format(date),
                            "Bonbu": bonbu,
                            "Center": center,
                            "Team": team,
                            "Guksa": guksa,
                            "Power Usage": total_power,
                        }
                    )
                else:
                    results.append(
                        {
                            "Customer Number": cust_no,
                            "Date": convert_date_format(date),
                            "Bonbu": bonbu,
                            "Center": center,
                            "Team": team,
                            "Guksa": guksa,
                            "Power Usage": None,
                        }
                    )
            else:
                results.append(
                    {
                        "Customer Number": cust_no,
                        "Date": convert_date_format(date),
                        "Bonbu": bonbu,
                        "Center": center,
                        "Team": team,
                        "Guksa": guksa,
                        "Power Usage": "API Error",
                    }
                )

        # Convert results to a DataFrame
        df = pd.DataFrame(results)

        # Generate timestamped Excel filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"kepco_daily_report_{timestamp}.xlsx"

        # Save DataFrame to Excel
        df.to_excel(excel_filename, index=False, engine="openpyxl")

        # Load the workbook and add a table
        wb = load_workbook(excel_filename)
        ws = wb.active
        # Define table range (from A1 to last column and row)
        table_range = f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}"
        table = Table(displayName="DailyReportTable", ref=table_range)
        # Apply table style
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(excel_filename)

        return excel_filename, True

        return excel_filename, True

    except Exception as e:
        print(f"Error generating daily report: {str(e)}")
        return None, False


def generate_kepco_15min_report():
    """Generate 15-minute interval power usage report (kepcoDailyData15min logic)."""
    try:
        # Set date to yesterday
        today = datetime.today() - timedelta(days=1)
        date = today.strftime("%Y%m%d")  # Format as YYYYMMDD

        # Input CSV file path
        input_csv = "/home/nit/kepco_report/kepcolist_gg.csv"

        # API URL and static parameters
        url = "https://opm.kepco.co.kr:11080/OpenAPI/getDayLpData.do"
        service_key = "bpb89eyd7bg430vckh8t"

        # Read customer numbers from CSV
        customer_data = pd.read_csv(input_csv, dtype=str)

        # Prepare a list to store the results
        results = []

        for _, row in customer_data.iterrows():
            cust_no = row.get("고객번호")
            if not cust_no:
                continue
            bonbu = row.get("본부명")
            center = row.get("센터")
            team = row.get("팀")
            guksa = row.get("국사")

            # API call
            params = {
                "custNo": cust_no,
                "date": date,
                "serviceKey": service_key,
                "returnType": "02",
            }
            response = requests.get(url, params=params)

            if response.status_code == 200:
                data = response.json()
                day_lp_data = data.get("dayLpDataInfoList", [])

                if day_lp_data:
                    for record in day_lp_data:
                        # Iterate over pwr_qtyXXXX keys
                        for key, value in record.items():
                            if key.startswith("pwr_qty") and isinstance(
                                value, (int, float)
                            ):
                                # Extract time from key (e.g., pwr_qty0015 → 00:15)
                                time_str = key[-4:]  # Last 4 characters (HHMM)
                                results.append(
                                    {
                                        "Customer Number": cust_no,
                                        "MeterNo": record.get("meterNo"),
                                        "Date": convert_date_format(date),
                                        "Time": time_str,
                                        "Bonbu": bonbu,
                                        "Center": center,
                                        "Team": team,
                                        "Guksa": guksa,
                                        "Power Usage": value,
                                    }
                                )
                else:
                    print(f"No data found for Customer Number {cust_no}")
                    continue
            else:
                print(
                    f"API Error for Customer Number {cust_no}: {response.status_code}"
                )
                continue

        # Convert results to a DataFrame
        df = pd.DataFrame(results)

        # Generate timestamped Excel filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"kepco_15min_report_{timestamp}.xlsx"

        # Save DataFrame to Excel
        df.to_excel(excel_filename, index=False, engine="openpyxl")

        # Load the workbook and add a table
        wb = load_workbook(excel_filename)
        ws = wb.active
        # Define table range (from A1 to last column and row)
        table_range = f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}"
        table = Table(displayName="Min15ReportTable", ref=table_range)
        # Apply table style
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(excel_filename)

        return excel_filename, True

    except Exception as e:
        print(f"Error generating 15min report: {str(e)}")
        return None, False


def send_email_with_attachments(excel_files):
    """Send email with multiple Excel file attachments."""
    try:
        # Email configuration
        smtp_server = "smtp.gmail.com"  # e.g., "smtp.gmail.com"
        smtp_port = 587  # e.g., 587 for TLS
        sender_email = "axgisuljiwontim@gmail.com"
        sender_password = "tjuugdecoipdjdyl"
        # recipient_email = "choi.js@kt.com"  # Replace with actual recipient email
        recipient_email = ["choi.js@kt.com", "h_w.kim@kt.com"]  # Replace with recipient emails


        # Create email
        msg = MIMEMultipart()
        msg["From"] = sender_email
        # msg["To"] = recipient_email
        msg['To'] = ", ".join(recipient_email)  # Join recipient emails with comma
        msg["Subject"] = (
            f"KEPCO 일일 및 15분 간격 보고서 - {datetime.now().strftime('%Y-%m-%d')}"
        )

        # Email body
        body = "첨부된 파일은 KEPCO 일일 전력 사용량 보고서와 15분 간격 전력 사용량 보고서입니다."
        msg.attach(MIMEText(body, "plain"))

        # Attach Excel files
        for excel_filename in excel_files:
            if excel_filename and os.path.exists(excel_filename):
                with open(excel_filename, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition", f"attachment; filename= {excel_filename}"
                )
                msg.attach(part)
            else:
                print(f"Attachment file not found: {excel_filename}")

        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)

        print(f"이메일 전송 성공: {', '.join(excel_files)}")
        return True

    except Exception as e:
        print(f"이메일 전송 오류: {str(e)}")
        return False


def main():
    # Generate daily report
    daily_excel, daily_success = generate_kepco_daily_report()

    # Generate 15-minute report
    min15_excel, min15_success = generate_kepco_15min_report()

    # Collect Excel files to attach
    excel_files = []
    if daily_success and daily_excel:
        excel_files.append(daily_excel)
    if min15_success and min15_excel:
        excel_files.append(min15_excel)

    # Send email with attachments
    if excel_files:
        email_success = send_email_with_attachments(excel_files)
        if email_success:
            print("보고서 생성 및 이메일 전송 성공.")
        else:
            print("보고서 생성 성공, 이메일 전송 실패.")

        # Clean up Excel files
        for excel_file in excel_files:
            try:
                os.remove(excel_file)
                print(f"임시 파일 삭제: {excel_file}")
            except Exception as e:
                print(f"임시 파일 삭제 오류: {str(e)}")
    else:
        print("보고서 생성 실패.")


if __name__ == "__main__":
    main()
